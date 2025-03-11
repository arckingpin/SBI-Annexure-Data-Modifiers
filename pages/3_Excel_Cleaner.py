import streamlit as st
import openpyxl
from io import BytesIO
import time
import pandas as pd
import re

st.set_page_config(page_title="Excel Cleaner", layout="centered")
st.title("Excel Cleaner")

uploaded_file = st.file_uploader("Select an Excel file", type=["xlsx"])
output_file_name = st.text_input("Enter a name for the processed file", "cleaned_file")

def highlight_changes(original):
    """Highlight only the leading and trailing whitespace."""
    if not original:
        return ""
    leading_match = re.match(r'^\s+', original)
    leading = leading_match.group(0) if leading_match else ""
    trailing_match = re.search(r'\s+$', original)
    trailing = trailing_match.group(0) if trailing_match else ""
    middle = original[len(leading): len(original) - len(trailing)] if trailing else original[len(leading):]
    highlighted = ""
    if leading:
        highlighted += f'<span style="background-color: yellow; color: #000">{leading}</span>'
    highlighted += middle
    if trailing:
        highlighted += f'<span style="background-color: yellow; color: #000">{trailing}</span>'
    return highlighted

# Container style: resizable with scroll bars
container_style = (
    "overflow: auto; resize: both; max-height: 400px; border: 1px solid #ccc; border-radius: 8px; "
    "padding: 15px; margin-bottom: 20px;"
)

# CSS style block to style our custom table and freeze the first row
frozen_rows_style = """
<style>
    table.custom-table {
        width: 100%;
        border-collapse: collapse;
        background-color: #2e2e2e;
        color: #fff;
    }
    /* Freeze only the first row in the header */
    table.custom-table thead tr {
        position: sticky;
        top: 0;
        background-color: #2e2e2e;
        z-index: 2;
    }
    table.custom-table th, table.custom-table td {
        border: 1px solid #555;
        padding: 4px;
        height: 30px;
        line-height: 30px;
    }
    /* Original table: preserve whitespace (line breaks, spaces) */
    table.original-table td {
        white-space: pre-wrap;
    }
    /* Processed table: no wrapping */
    table.processed-table td {
        white-space: nowrap;
    }
</style>
"""

def df_to_html_with_frozen_header(df, table_class):
    """
    Generate HTML for a table with the first row wrapped in <thead> (for freezing)
    and the remaining rows in <tbody>.
    """
    html = f'<table class="{table_class}">'
    html += "<thead>"
    header_rows = df.iloc[:1]
    for _, row in header_rows.iterrows():
        html += "<tr>"
        for cell in row:
            html += f"<th>{cell}</th>"
        html += "</tr>"
    html += "</thead>"
    html += "<tbody>"
    if len(df) > 1:
        body_rows = df.iloc[1:]
        for _, row in body_rows.iterrows():
            html += "<tr>"
            for cell in row:
                html += f"<td>{cell}</td>"
            html += "</tr>"
    html += "</tbody></table>"
    return html

# -------- ORIGINAL DATA PREVIEW --------
if uploaded_file:
    try:
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
        sheet = workbook.active
        data_original = []

        for row in sheet.iter_rows():
            original_row = []
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                original_row.append(highlight_changes(val))
            data_original.append(original_row)
        df_original = pd.DataFrame(data_original)
        html_original = df_to_html_with_frozen_header(df_original, "custom-table original-table")
        html_original = frozen_rows_style + html_original

        st.subheader("Original Data (Highlighted Leading/Trailing Spaces)")
        st.markdown(f'<div style="{container_style}">{html_original}</div>', unsafe_allow_html=True)
    except Exception as e:
        st.error(f"An error occurred while loading the file: {e}")

# -------- CLEANING ACTION --------
if st.button("Start Cleaning"):
    if uploaded_file is None:
        st.warning("Please upload an Excel file!")
    elif not output_file_name:
        st.warning("Please enter a name for the processed file!")
    else:
        try:
            workbook = openpyxl.load_workbook(uploaded_file)
            sheet = workbook.active
            data_cleaned = []
            total_cells = sheet.max_row * sheet.max_column
            processed_cells = 0

            progress_bar = st.progress(0)
            progress_text = st.empty()

            for row in sheet.iter_rows():
                cleaned_row = []
                for cell in row:
                    val = str(cell.value) if cell.value is not None else ""
                    cleaned_value = val.strip()  # Remove leading/trailing spaces
                    cleaned_row.append(cleaned_value)
                    cell.value = cleaned_value
                    processed_cells += 1
                    progress_bar.progress(int((processed_cells / total_cells) * 100))
                    progress_text.text(f"Progress: {int((processed_cells / total_cells) * 100)}%")
                    time.sleep(0.001)
                data_cleaned.append(cleaned_row)
            df_cleaned = pd.DataFrame(data_cleaned)
            html_cleaned = df_to_html_with_frozen_header(df_cleaned, "custom-table processed-table")
            html_cleaned = frozen_rows_style + html_cleaned

            st.subheader("Cleaned Data")
            st.markdown(f'<div style="{container_style}">{html_cleaned}</div>', unsafe_allow_html=True)

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            st.success("File cleaned successfully!")
            output_filename = f"{uploaded_file.name.replace('.xlsx', '')}_{output_file_name}.xlsx"
            st.download_button(
                label="Download Cleaned File",
                data=output,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"An error occurred: {e}")
